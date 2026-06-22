import os
import re
import json
import logging
import requests
from datetime import date
from pathlib import Path
from urllib.parse import urljoin

from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font


load_dotenv()

# =========================
# 基础配置
# =========================

EXPORT_DIR = Path("data/company_results")
EXPORT_DIR.mkdir(parents=True, exist_ok=True)

OPEN_CORPORATES_BASE_URL = "https://opencorporates.com/"

# 系统A地址，例如：
# OPEN_CORPORATES_API_BASE_URL=http://127.0.0.1:8081
SYSTEM_A_BASE_URL = os.getenv("OPEN_CORPORATES_API_BASE_URL", "").rstrip("/")

REQUEST_TIMEOUT = int(os.getenv("OPEN_CORPORATES_REQUEST_TIMEOUT", "30"))


# =========================
# XLSX 字段配置
# =========================

XLSX_FIELDNAMES = [
    "query_name",
    "final_status",
    "search_status",
    "company_name",
    "company_address",
    "company_start_date",
    "company_status",
    "company_href",
]

XLSX_HEADER_LABELS = {
    "query_name": "待查list_name",
    "final_status": "最终判定",
    "search_status": "检索状态",
    "company_name": "检索到的公司名称",
    "company_address": "注册地址",
    "company_start_date": "注册时间",
    "company_status": "inactivate",
    "company_href": "检索网站地址",
}

FINAL_STATUS_WHITELIST = "白名单/有效公司"
FINAL_STATUS_BLACKLIST = "黑名单/无效公司"
FINAL_STATUS_PENDING = "待定，请重试/系统原因"

MONTH_NUMBERS = {
    "jan": 1,
    "january": 1,
    "feb": 2,
    "february": 2,
    "mar": 3,
    "march": 3,
    "apr": 4,
    "april": 4,
    "may": 5,
    "jun": 6,
    "june": 6,
    "jul": 7,
    "july": 7,
    "aug": 8,
    "august": 8,
    "sep": 9,
    "sept": 9,
    "september": 9,
    "oct": 10,
    "october": 10,
    "nov": 11,
    "november": 11,
    "dec": 12,
    "december": 12,
}


# =========================
# 系统A接口调用
# =========================

def _get_system_a_base_url() -> str:
    if not SYSTEM_A_BASE_URL:
        raise ValueError("缺少环境变量 OPEN_CORPORATES_API_BASE_URL")
    return SYSTEM_A_BASE_URL


def submit_open_corporates_task(company_name_list: list[str], search_name: str) -> dict:
    """
    提交 OpenCorporates 异步任务到系统A。

    调用系统A：
    POST /openCorporates/search/async

    返回示例：
    {
        "success": true,
        "task_id": "...",
        "message": "OpenCorporates查询任务已创建"
    }
    """
    base_url = _get_system_a_base_url()

    response = requests.post(
        f"{base_url}/openCorporates/search/async",
        json={
            "company_name_list": company_name_list,
            "search_name": search_name,
        },
        timeout=REQUEST_TIMEOUT,
    )
    response.raise_for_status()

    payload = response.json()
    task_id = payload.get("task_id")

    if not task_id:
        raise ValueError(f"系统A未返回 task_id: {payload}")

    return payload


def get_open_corporates_task(task_id: str) -> dict:
    """
    查询系统A中的单个 OpenCorporates 任务详情。

    调用系统A：
    GET /openCorporates/task/{task_id}
    """
    base_url = _get_system_a_base_url()

    response = requests.get(
        f"{base_url}/openCorporates/company-results/{task_id}",
        timeout=REQUEST_TIMEOUT,
    )
    response.raise_for_status()
    return response.json()


def retry_open_corporates_task(task_id: str) -> dict:
    """
    重试 OpenCorporates 任务：读取原任务参数后重新提交，并删除原任务。

    系统A未提供专用 retry 接口，因此等价于：
    GET /openCorporates/task/{task_id}
    POST /openCorporates/search/async
    DELETE /tasks/open-corporates/{task_id}
    """
    base_url = _get_system_a_base_url()

    response = requests.get(
        f"{base_url}/openCorporates/task/{task_id}",
        timeout=REQUEST_TIMEOUT,
    )
    response.raise_for_status()
    task = response.json()

    input_data = task.get("input_data") or {}
    company_name_list = normalize_company_name_list(input_data.get("company_name_list"))
    search_name = str(
        input_data.get("search_name") or task.get("search_name") or ""
    ).strip()

    if not company_name_list:
        raise ValueError("原任务缺少 company_name_list，无法重试")

    if not search_name:
        raise ValueError("原任务缺少 search_name，无法重试")

    result = submit_open_corporates_task(company_name_list, search_name)
    result["source_task_id"] = task_id

    try:
        delete_open_corporates_task(task_id)
        result["source_task_deleted"] = True
    except requests.HTTPError as e:
        status_code = e.response.status_code if e.response is not None else None
        if status_code == 404:
            result["source_task_deleted"] = True
        else:
            result["source_task_deleted"] = False
            result["source_task_delete_error"] = str(e)
    except requests.Timeout:
        result["source_task_deleted"] = False
        result["source_task_delete_error"] = "删除原任务超时"

    return result


def delete_open_corporates_task(task_id: str) -> dict:
    """
    删除系统A中的 OpenCorporates 任务。

    调用系统A:
    DELETE /tasks/open-corporates/{task_id}
    """
    base_url = _get_system_a_base_url()

    response = requests.delete(
        f"{base_url}/tasks/open-corporates/{task_id}",
        headers={"accept": "application/json"},
        timeout=REQUEST_TIMEOUT,
    )
    response.raise_for_status()

    if response.content:
        try:
            return response.json()
        except ValueError:
            return {"message": response.text or "删除成功"}

    return {"message": "删除成功"}


def list_open_corporates_tasks(limit: int = 100, offset: int = 0) -> dict:
    """
    查询系统A中的 OpenCorporates 历史任务列表。

    调用系统A：
    GET /openCorporates/tasks?limit=100&offset=0
    """
    base_url = _get_system_a_base_url()

    response = requests.get(
        f"{base_url}/openCorporates/tasks",
        params={
            "limit": limit,
            "offset": offset,
        },
        timeout=REQUEST_TIMEOUT,
    )
    response.raise_for_status()
    return response.json()


# =========================
# 输入清洗
# =========================

def clean_company_names(text: str) -> list[str]:
    """
    清洗用户输入的公司名。
    支持用户一行一个公司名。
    """
    result = []

    for line in text.splitlines():
        line = line.strip()
        line = re.sub(r'^"+|"+$', "", line)
        line = line.strip()
        line = re.sub(r"\s+", " ", line)

        if not line:
            continue

        result.append(line)

    return result


def normalize_company_name_list(company_name_list) -> list[str]:
    """
    清洗 API 传入的公司名数组。
    """
    if not isinstance(company_name_list, list):
        return []

    return [
        str(name).strip()
        for name in company_name_list
        if isinstance(name, str) and name.strip()
    ]


# =========================
# 系统A结果解析
# =========================

def _decode_json_value(value):
    """
    递归解析 JSON 字符串，直到得到 Python 对象。

    兼容这些情况：
    1. dict
    2. JSON string
    3. JSON string inside JSON string
    """
    while isinstance(value, str):
        text = value.strip()
        if not text:
            return {}
        value = json.loads(text)
    return value

def extract_company_result_from_task(task_payload: dict) -> dict:
    """
    从任务详情中提取 OpenCorporates 检索结果。

    输入格式：
    {
        "status": "completed",
        "company_results": [
            {"company_name": "TPD ENTERPRISE INC", "result": [...]},
            ...
        ]
    }

    返回：
    {
        "TPD ENTERPRISE INC": [...],
        ...
    }
    """
    if not isinstance(task_payload, dict):
        raise ValueError("任务详情格式错误")

    status = task_payload.get("status")
    if status != "completed":
        raise ValueError(f"任务尚未完成，当前状态: {status}")

    company_results = task_payload.get("company_results") or []
    if not isinstance(company_results, list):
        raise ValueError("company_results 格式错误，应为列表")

    return {
        item["company_name"]: item.get("result")
        for item in company_results
        if isinstance(item, dict) and "company_name" in item
    }



# =========================
# 结果清洗与判定
# =========================

def _looks_like_failure(record: dict) -> str:
    """
    判断单条接口记录是否表示失败。
    如果失败，返回错误信息；否则返回空字符串。
    """
    status = str(record.get("status") or record.get("search_status") or "").strip().lower()

    if status in {"error", "failed", "fail", "失败", "检索失败"}:
        return str(record.get("message") or record.get("error") or "检索失败").strip()

    if record.get("success") is False:
        return str(record.get("message") or record.get("error") or "检索失败").strip()

    if record.get("error"):
        return str(record.get("error")).strip()

    company_status = str(record.get("company_status") or "").strip()
    if "搜索结果加载失败或未找到结果" in company_status:
        return company_status

    return ""


def _empty_row(query_name: str, search_status: str, final_status: str = "") -> dict:
    """
    生成一行占位结果。
    """
    return {
        "query_name": query_name,
        "final_status": final_status,
        "search_status": search_status,
    }


def _normalize_company_name(name) -> str:
    """
    归一化公司名：去掉常见符号和空格，并统一转为大写。
    """
    chars_to_remove = " -.,'&"
    return str(name or "").translate(
        str.maketrans("", "", chars_to_remove)
    ).upper()


def _is_same_company_name(query_name: str, company_name) -> bool:
    """
    判断待查公司名和检索结果公司名是否完全一致。
    """
    normalized_query = _normalize_company_name(query_name)
    normalized_company = _normalize_company_name(company_name)

    return bool(
        normalized_query
        and normalized_company
        and normalized_query == normalized_company
    )


def normalize_company_href(href) -> str:
    """
    补全 OpenCorporates 公司链接。
    """
    text = str(href or "").strip()

    if not text:
        return ""

    lower_text = text.lower()

    if lower_text.startswith(("http://", "https://")):
        return text

    if text.startswith("//"):
        return f"https:{text}"

    if lower_text.startswith("opencorporates.com/"):
        return f"https://{text}"

    return urljoin(OPEN_CORPORATES_BASE_URL, text)


def _is_inactive_company_status(company_status) -> bool:
    """
    判断公司状态是否为 inactive / inactivate。
    """
    status = str(company_status or "").strip().lower()

    return (
        status in {"inactive", "inactivate"}
        or status.startswith(("inactive ", "inactivate "))
    )


def _normalize_year(year_text: str) -> int:
    year = int(year_text)

    if year < 100:
        return 2000 + year if year <= 68 else 1900 + year

    return year


def _parse_start_date(company_start_date) -> date | None:
    """
    解析 OpenCorporates 常见注册时间格式，用于多结果合并时取最早记录。
    """
    text = str(company_start_date or "").strip()

    if not text:
        return None

    parts = [
        part.strip(".,")
        for part in text.replace("-", " ").replace("/", " ").split()
        if part.strip(".,")
    ]

    if len(parts) < 3:
        return None

    try:
        if parts[0].isdigit() and parts[1].lower() in MONTH_NUMBERS:
            day = int(parts[0])
            month = MONTH_NUMBERS[parts[1].lower()]
            year = _normalize_year(parts[2])

        elif parts[0].isdigit() and len(parts[0]) == 4 and parts[1].lower() in MONTH_NUMBERS:
            year = _normalize_year(parts[0])
            month = MONTH_NUMBERS[parts[1].lower()]
            day = int(parts[2])

        elif parts[0].isdigit() and len(parts[0]) == 4 and parts[1].isdigit() and parts[2].isdigit():
            year = _normalize_year(parts[0])
            month = int(parts[1])
            day = int(parts[2])

        else:
            return None

        return date(year, month, day)

    except (TypeError, ValueError):
        return None


def _select_final_record(records: list[dict]) -> dict:
    """
    多条命中结果合并为一条：
    1. 优先选择有地址的记录
    2. 再选择注册时间最早的记录
    """
    records_with_address = [
        record
        for record in records
        if str(record.get("company_address") or "").strip()
    ]

    candidates = records_with_address or records

    def sort_key(indexed_record):
        index, record = indexed_record
        start_date = _parse_start_date(record.get("company_start_date"))

        return (
            0 if start_date else 1,
            start_date or date.max,
            index,
        )

    return min(enumerate(candidates), key=sort_key)[1]


def _final_row_from_record(query_name: str, record: dict, final_status: str) -> dict:
    """
    把命中的公司记录收敛为最终 XLSX 的单行结果。
    """
    row = dict(record)

    row["query_name"] = query_name
    row["final_status"] = final_status
    row["search_status"] = row.get("search_status") or "检索成功"
    row["company_href"] = normalize_company_href(row.get("company_href"))

    return row


def _rows_for_query(query_name: str, records) -> tuple[list[dict], str]:
    """
    清洗单个 query_name 的检索记录。
    返回：
    (
        rows,
        query_status
    )

    query_status 只会是：
    success / failed / no_result
    """
    if records is None or records == []:
        return [_empty_row(query_name, "检索无结果", FINAL_STATUS_BLACKLIST)], "no_result"

    if isinstance(records, str):
        return [_empty_row(query_name, f"检索失败: {records}", FINAL_STATUS_PENDING)], "failed"

    if isinstance(records, dict):
        failure_message = _looks_like_failure(records)

        if failure_message:
            return [_empty_row(query_name, f"检索失败: {failure_message}", FINAL_STATUS_PENDING)], "failed"

        records = [records]

    if not isinstance(records, list):
        return [_empty_row(query_name, "结果格式异常", FINAL_STATUS_PENDING)], "failed"
    

    matched_records = []
    failure_messages = []

    for record in records:
        if not isinstance(record, dict):
            failure_messages.append("结果格式异常")
            continue

        failure_message = _looks_like_failure(record)

        if failure_message:
            failure_messages.append(failure_message)
            continue

        if not _is_same_company_name(query_name, record.get("company_name")):
            continue

        matched_records.append(record)

    if matched_records:
        active_records = [
            record
            for record in matched_records
            if not _is_inactive_company_status(record.get("company_status"))
        ]

        if active_records:
            return [
                _final_row_from_record(
                    query_name,
                    _select_final_record(active_records),
                    FINAL_STATUS_WHITELIST,
                )
            ], "success"

        return [
            _final_row_from_record(
                query_name,
                _select_final_record(matched_records),
                FINAL_STATUS_BLACKLIST,
            )
        ], "success"

    if failure_messages:
        message = "; ".join(
            dict.fromkeys(
                str(msg).strip()
                for msg in failure_messages
                if str(msg).strip()
            )
        )

        return [
            _empty_row(
                query_name,
                f"检索失败: {message or '系统原因'}",
                FINAL_STATUS_PENDING,
            )
        ], "failed"

    return [_empty_row(query_name, "检索无精确匹配结果", FINAL_STATUS_BLACKLIST)], "no_result"


def _build_summary(total: int, counts: dict) -> dict:
    """
    生成前端展示用汇总。
    """
    failed = counts["failed"]
    no_result = counts["no_result"]
    success = counts["success"]

    return {
        "total": total,
        "success": success,
        "failed": failed,
        "no_result": no_result,
        "has_partial_failure": failed > 0 and success > 0,
        "has_failure": failed > 0,
        "has_no_result": no_result > 0,
    }


def build_result_rows(data: dict, query_names: list[str] | None = None) -> tuple[list[dict], dict]:
    """
    把系统A原始结果清洗成可写入 XLSX 的 rows。
    """
    counts = {
        "success": 0,
        "failed": 0,
        "no_result": 0,
    }

    rows = []

    ordered_query_names = list(query_names or data.keys())
    ordered_query_names.extend(
        key
        for key in data.keys()
        if key not in ordered_query_names
    )

    for query_name in ordered_query_names:
        query_rows, query_status = _rows_for_query(
            query_name,
            data.get(query_name),
        )

        counts[query_status] += 1
        rows.extend(query_rows)

    return rows, _build_summary(len(ordered_query_names), counts)


def group_result_rows(rows: list[dict]) -> dict:
    """
    按 query_name 分组，方便前端展示。
    """
    grouped = {}

    for row in rows:
        name = row.get("query_name") or row.get("company_name") or "unknown"
        grouped.setdefault(name, []).append(row)

    return grouped


# =========================
# XLSX 生成
# =========================

def save_to_xlsx(
    data: dict,
    task_id: str,
    query_names: list[str] | None = None,
) -> tuple[Path, dict, dict]:
    """
    将 OpenCorporates 结果清洗后保存为 XLSX。

    返回：
    (
        xlsx_path,
        summary,
        grouped_result_data
    )
    """
    xlsx_path = EXPORT_DIR / f"{task_id}.xlsx"

    rows, summary = build_result_rows(data, query_names)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "company_results"

    for col_index, field in enumerate(XLSX_FIELDNAMES, start=1):
        cell = sheet.cell(
            row=1,
            column=col_index,
            value=XLSX_HEADER_LABELS.get(field, field),
        )
        cell.font = Font(bold=True)
        sheet.column_dimensions[cell.column_letter].width = 24

    for row_index, row in enumerate(rows, start=2):
        for col_index, field in enumerate(XLSX_FIELDNAMES, start=1):
            sheet.cell(
                row=row_index,
                column=col_index,
                value=str(row.get(field, "") or ""),
            )

    workbook.save(xlsx_path)

    return xlsx_path, summary, group_result_rows(rows)


def save_failure_xlsx(
    task_id: str,
    query_names: list[str],
    message: str,
) -> tuple[Path, dict, dict]:
    """
    整批请求失败时生成失败明细 XLSX。
    """
    data = {
        name: {
            "status": "error",
            "message": message,
        }
        for name in query_names
    }

    return save_to_xlsx(data, task_id, query_names)


def create_xlsx_from_open_corporates_task(task_id: str) -> tuple[Path, dict, dict]:
    """
    根据系统A的 task_id 查询结果，并生成 XLSX。

    适合 Flask 下载接口中调用。
    """
    task_payload = get_open_corporates_task(task_id)

    input_data = task_payload.get("input_data") or {}
    query_names = input_data.get("company_name_list") or []

    company_result = extract_company_result_from_task(task_payload)

    return save_to_xlsx(
        company_result,
        task_id,
        query_names,
    )
